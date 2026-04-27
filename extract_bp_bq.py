#!/usr/bin/env python3
from __future__ import annotations
"""
Extrai dados de Real vs Orçado (BP) da planilha Excel
e escreve no BigQuery (tabela orcamento_dre).

Adaptado de extract_orcamento.py — mantém lógica de detecção e DRE_MAP,
substitui output JSON por BigQuery.

Variáveis de ambiente:
  GCP_PROJECT_ID                 — projeto GCP
  BQ_DATASET                     — dataset BigQuery (default: studio_koti)
  GOOGLE_APPLICATION_CREDENTIALS — path para JSON da service account

Uso:
  pip install openpyxl google-cloud-bigquery db-dtypes
  python extract_bp_bq.py [caminho_planilha]
"""

import os
import sys
import glob
import unicodedata
from datetime import datetime

from google.cloud import bigquery

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ICLOUD_BP_DIR = os.path.expanduser(
    "~/Library/Mobile Documents/com~apple~CloudDocs/Studio Koti/Dashboard/Dados BP"
)
GCP_PROJECT_ID = os.environ.get("GCP_PROJECT_ID", "")
BQ_DATASET = os.environ.get("BQ_DATASET", "studio_koti")


def find_bp_file() -> str | None:
    """Busca arquivo BP no diretório do script e na pasta iCloud da Koti."""
    candidates = [
        os.path.join(SCRIPT_DIR, "BP.xlsx"),
        os.path.join(ICLOUD_BP_DIR, "BP.xlsx"),
    ]
    for bp in candidates:
        if os.path.exists(bp):
            return bp

    matches = glob.glob(os.path.join(SCRIPT_DIR, "BP*.xlsx"))
    matches += glob.glob(os.path.join(ICLOUD_BP_DIR, "BP*.xlsx"))
    if matches:
        return sorted(matches)[-1]
    return None


# Mapeamento das linhas do DRE
# (real_row, bp_row, label, section, level)
# level: 0=total, 1=subtotal, 2=detalhe
DRE_MAP = [
    (26, 51, "Receita Bruta", "receita", 0),
    (35, 60, "Impostos", "impostos", 0),
    (44, 68, "Receita Líquida", "receita_liq", 0),
    (46, 70, "Custos Operacionais", "custos", 0),
    (75, 98, "Margem de Contribuição", "margem", 0),
    (78, 101, "Despesas Gerais e Adm", "sga", 0),
    (80, 103, "Salários e Encargos", "sga", 1),
    (147, 168, "EBITDA", "ebitda", 0),
    (150, 171, "Receitas/Despesas Financeiras", "financeiro", 1),
    ((41, 42), (236, 238), "IRPJ/CSLL", "impostos_renda", 0),
    (169, 186, "Lucro Líquido", "ll", 0),
]
MONTH_COLS = {f"2026-{m:02d}": 3 + m for m in range(1, 13)}


def _norm_label(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.replace("ç", "c")


def read_val(ws, row: int | tuple[int, ...] | None, col: int) -> float:
    """Lê valor numérico de uma célula."""
    if row is None:
        return 0
    if isinstance(row, tuple):
        return round(sum(read_val(ws, r, col) for r in row), 2)
    v = ws.cell(row=row, column=col).value
    return round(v, 2) if isinstance(v, (int, float)) else 0


def validate_dre_map(ws_real, ws_bp) -> int:
    """Valida que os labels do DRE_MAP correspondem às células da planilha.
    Retorna o número de mismatches encontrados."""
    # Labels esperados na coluna A (ou B) de cada aba, indexados por row
    mismatches = 0
    for real_row, bp_row, label, section, level in DRE_MAP:
        for ws, row, aba in [(ws_real, real_row, "Realizado"), (ws_bp, bp_row, "BP")]:
            if row is None:
                continue
            if isinstance(row, tuple):
                continue
            # Label geralmente está na coluna 2 (B) ou 3 (C), testar ambas
            cell_b = ws.cell(row=row, column=2).value
            cell_c = ws.cell(row=row, column=3).value
            cell_val = str(cell_b or "").strip() if cell_b else str(cell_c or "").strip()
            found = _norm_label(cell_val)
            expected = _norm_label(label)
            aliases = {
                "impostos": ["imposto retido na fonte", "impostos s/ receita"],
                "receitas/despesas financeiras": ["receitas e despesas financeiras"],
                "lucro liquido": ["ll"],
            }
            accepted = [expected] + aliases.get(expected, [])
            if cell_val and not any(a in found or found in a for a in accepted):
                print(f"  ⚠ WARNING: {aba} row {row} esperado '{label}', encontrado '{cell_val}'")
                mismatches += 1
    return mismatches


def extract_to_bq(filepath: str) -> None:
    """Extrai dados da planilha BP e carrega no BigQuery."""
    from openpyxl import load_workbook

    print(f"📊 Lendo planilha: {filepath}")
    wb = load_workbook(filepath, data_only=True)

    ws_real = wb["Realizado"]
    ws_bp = wb["BP"]

    # Validar DRE_MAP contra labels da planilha
    mismatches = validate_dre_map(ws_real, ws_bp)
    if mismatches > 3:
        print(f"  ❌ ERRO: {mismatches} labels não batem com a planilha. Abortando.")
        sys.exit(1)
    elif mismatches > 0:
        print(f"  ⚠ {mismatches} label(s) com divergência — continuando com cautela")

    # Detectar meses com dados reais (Receita Bruta != 0)
    meses_com_real: set[str] = set()
    for mes, col in sorted(MONTH_COLS.items()):
        rb = read_val(ws_real, 26, col)
        if rb != 0:
            meses_com_real.add(mes)

    print(f"  Meses com dados reais: {', '.join(sorted(meses_com_real)) if meses_com_real else 'nenhum'}")

    sync_ts = datetime.utcnow().isoformat()

    # Flatten: 1 linha por item × mês
    rows: list[dict] = []
    for real_row, bp_row, label, section, level in DRE_MAP:
        for mes, col in sorted(MONTH_COLS.items()):
            valor_real = read_val(ws_real, real_row, col)
            valor_bp = read_val(ws_bp, bp_row, col)

            # variacao_pct: (real - bp) / abs(bp) * 100, NULL se bp == 0
            variacao_pct = None
            if valor_bp != 0:
                variacao_pct = round((valor_real - valor_bp) / abs(valor_bp) * 100, 2)

            rows.append({
                "label": label,
                "section": section,
                "level": level,
                "mes": mes,
                "valor_real": valor_real,
                "valor_bp": valor_bp,
                "variacao_pct": variacao_pct,
                "mes_com_real": mes in meses_com_real,
                "sync_timestamp": sync_ts,
            })

    if not rows:
        print("  ⚠ Nenhuma linha extraída — pulando")
        return

    # Carregar no BigQuery
    print(f"\n📤 Carregando {len(rows)} linhas no BigQuery...")
    client = bigquery.Client(project=GCP_PROJECT_ID)
    table = f"{GCP_PROJECT_ID}.{BQ_DATASET}.orcamento_dre"

    job_config = bigquery.LoadJobConfig(
        write_disposition="WRITE_TRUNCATE",
        source_format=bigquery.SourceFormat.NEWLINE_DELIMITED_JSON,
        autodetect=False,
    )

    job = client.load_table_from_json(rows, table, job_config=job_config)
    job.result()

    print(f"  ✅ {len(rows)} linhas carregadas em orcamento_dre (WRITE_TRUNCATE)")
    print(f"  ✅ {len(DRE_MAP)} linhas DRE × {len(MONTH_COLS)} meses")


def main() -> None:
    if not GCP_PROJECT_ID:
        print("❌ Configure GCP_PROJECT_ID!")
        sys.exit(1)

    filepath = sys.argv[1] if len(sys.argv) > 1 else find_bp_file()

    if not filepath or not os.path.exists(filepath):
        print("⚠ Planilha BP não encontrada — orcamento_dre não será atualizado")
        sys.exit(0)  # Exit 0 para não quebrar o workflow

    extract_to_bq(filepath)


if __name__ == "__main__":
    main()
