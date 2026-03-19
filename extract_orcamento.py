#!/usr/bin/env python3
"""
Extrai dados de Real vs Orçado (BP) da planilha Excel
e gera dados_orcamento.json para consumo pelo dashboard.

Uso:
  pip install openpyxl
  python3 extract_orcamento.py [caminho_planilha]

Se nenhum caminho for fornecido, busca BP.xlsx no mesmo diretório.
"""

import json
import os
import sys
import glob

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Encontrar planilha BP
def find_bp_file():
    """Busca arquivo BP no diretório do script."""
    # Tenta BP.xlsx primeiro
    bp = os.path.join(SCRIPT_DIR, "BP.xlsx")
    if os.path.exists(bp):
        return bp
    # Tenta qualquer arquivo BP*.xlsx
    matches = glob.glob(os.path.join(SCRIPT_DIR, "BP*.xlsx"))
    if matches:
        return sorted(matches)[-1]  # Pega o mais recente por nome
    return None


# Mapeamento das linhas do DRE
# (real_row, bp_row, label, section, level)
# level: 0=total, 1=subtotal, 2=detalhe
DRE_MAP = [
    (26, 26, "Receita Bruta", "receita", 0),
    (27, None, "SK", "receita", 2),
    (28, 29, "BK", "receita", 2),
    (29, 30, "RT", "receita", 2),
    (30, None, "Aditivo", "receita", 2),
    (None, 32, "Vendas RP", "receita", 2),
    (35, 35, "Impostos", "impostos", 0),
    (37, 37, "ICMS", "impostos", 2),
    (38, 38, "Crédito de ICMS", "impostos", 2),
    (39, 39, "ISS", "impostos", 2),
    (40, 40, "PIS/COFINS", "impostos", 2),
    (42, 42, "Receita Líquida", "receita_liq", 0),
    (44, 44, "Custos Operacionais", "custos", 0),
    (45, 45, "Comissões Externas", "custos", 2),
    (47, 46, "Comissões Internas", "custos", 2),
    (48, 47, "Obras (Total)", "custos", 1),
    (73, 71, "Margem de Contribuição", "margem", 0),
    (76, 74, "Despesas Gerais e Adm", "sga", 0),
    (78, 76, "Salários e Encargos", "sga", 1),
    (92, 90, "Despesas Administrativas", "sga", 1),
    (116, 113, "Despesas Comerciais", "sga", 1),
    (121, 117, "Despesas com Imóvel", "sga", 1),
    (131, 127, "Despesas com Veículos", "sga", 1),
    (137, 133, "Despesas com Diretoria", "sga", 1),
    (145, 141, "EBITDA", "ebitda", 0),
    (148, 144, "Receitas/Despesas Financeiras", "financeiro", 1),
    (163, 159, "IRPJ/CSLL", "impostos_renda", 0),
    (168, 164, "Lucro Líquido", "ll", 0),
]

MONTH_COLS = {f"2026-{m:02d}": 3 + m for m in range(1, 13)}


def read_val(ws, row, col):
    if row is None:
        return 0
    v = ws.cell(row=row, column=col).value
    return round(v, 2) if isinstance(v, (int, float)) else 0


def extract(filepath):
    from openpyxl import load_workbook

    print(f"📊 Lendo planilha: {filepath}")
    wb = load_workbook(filepath, data_only=True)

    ws_real = wb['Realizado']
    ws_bp = wb['BP']

    # Detectar meses com dados reais (Receita Bruta != 0)
    meses_com_real = []
    for mes, col in sorted(MONTH_COLS.items()):
        rb = read_val(ws_real, 26, col)
        if rb != 0:
            meses_com_real.append(mes)

    print(f"  Meses com dados reais: {', '.join(meses_com_real) if meses_com_real else 'nenhum'}")

    # Extrair dados
    orcamento = {
        "meses_disponiveis": list(sorted(MONTH_COLS.keys())),
        "meses_com_real": meses_com_real,
        "dre": []
    }

    for real_row, bp_row, label, section, level in DRE_MAP:
        line = {
            "label": label,
            "section": section,
            "level": level,
            "bp": {},
            "real": {},
        }
        for mes, col in sorted(MONTH_COLS.items()):
            line["bp"][mes] = read_val(ws_bp, bp_row, col)
            line["real"][mes] = read_val(ws_real, real_row, col)
        orcamento["dre"].append(line)

    return orcamento


def main():
    filepath = sys.argv[1] if len(sys.argv) > 1 else find_bp_file()

    if not filepath or not os.path.exists(filepath):
        print("⚠ Planilha BP não encontrada — dados_orcamento.json não será atualizado")
        sys.exit(0)  # Exit 0 para não quebrar o workflow

    orcamento = extract(filepath)

    output = os.path.join(SCRIPT_DIR, "dados_orcamento.json")
    with open(output, "w", encoding="utf-8") as f:
        json.dump(orcamento, f, ensure_ascii=False, indent=2)

    print(f"  ✅ {len(orcamento['dre'])} linhas DRE extraídas")
    print(f"  ✅ Salvo em: {output} ({os.path.getsize(output) / 1024:.1f} KB)")


if __name__ == "__main__":
    main()
