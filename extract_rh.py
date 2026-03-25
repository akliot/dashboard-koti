#!/usr/bin/env python3
"""
Extrai dados AGREGADOS da folha de pagamentos para o dashboard.
Gera rh_data.json com totais por área/rubrica — SEM dados pessoais (LGPD).

Uso:
  python3 extract_rh.py [caminho_planilha]
  python3 extract_rh.py  # busca em ~/Downloads/Folha de Pagamentos 2026.xlsx

NUNCA expor: nomes, CPF, salário individual, dados bancários.
"""

import json
import os
import sys
from datetime import date
from collections import defaultdict

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_PATH = os.path.expanduser("~/Downloads/Folha de Pagamentos 2026.xlsx")

# Meses e colunas no RESUMO (Col 2=Jan, 3=Fev, ..., 13=Dez)
MESES_COLS = {f"2026-{m:02d}": m + 1 for m in range(1, 13)}
MESES_NOMES = ["Janeiro 26", "Fevereiro 26", "Março 26", "Abril 26",
               "Maio 26", "Junho 26", "Julho 26", "Agosto 26",
               "Setembro 26", "Outubro 26", "Novembro 26", "Dezembro 26"]

# Rubricas no RESUMO (linhas 4-12)
RUBRICAS = [
    (4, "salarios", "Salário Base"),
    (5, "softwares", "Softwares"),
    (6, "beneficios", "Benefícios"),
    (7, "comissao", "Comissão"),
    (8, "rescisao", "Rescisão"),
    (9, "bonus", "Bônus"),
    (10, "impostos", "Impostos/Encargos"),
    (11, "treinamentos", "Treinamentos"),
    (12, "decimo_terceiro", "13º Salário"),
]

# HC no RESUMO (linhas 15-20)
HC_ROWS = {
    "inicio": 15,
    "saida_vol": 16,
    "saida_inv": 17,
    "admissoes": 18,
    "final": 19,
    "turnover": 20,
}

# Áreas no RESUMO (linhas 23-31)
AREAS_START = 23
AREAS_END = 31


def read_val(ws, row, col):
    v = ws.cell(row=row, column=col).value
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return round(v, 2)
    return 0


def detect_status(wb):
    """Detecta status de cada mês: realizado, andamento, projecao."""
    hoje = date.today()
    mes_atual = f"2026-{hoje.month:02d}"
    status = {}
    for mes_key in MESES_COLS:
        m = int(mes_key.split("-")[1])
        if mes_key < mes_atual:
            status[mes_key] = "realizado"
        elif mes_key == mes_atual:
            status[mes_key] = "andamento"
        else:
            status[mes_key] = "projecao"
    return status, mes_atual


def extract_resumo(wb):
    """Extrai dados da aba RESUMO (agregados — sem dados pessoais)."""
    ws = wb["RESUMO"]

    # Custo mensal por rubrica
    custo_mensal = {}
    for mes_key, col in MESES_COLS.items():
        mes_data = {}
        total = 0
        for row, key, _ in RUBRICAS:
            v = read_val(ws, row, col)
            mes_data[key] = v
            total += v
        mes_data["total"] = round(read_val(ws, 13, col), 2)
        custo_mensal[mes_key] = mes_data

    # HC
    hc = {}
    for mes_key, col in MESES_COLS.items():
        hc[mes_key] = {
            "inicio": int(read_val(ws, HC_ROWS["inicio"], col)),
            "admissoes": int(read_val(ws, HC_ROWS["admissoes"], col)),
            "saida_vol": int(read_val(ws, HC_ROWS["saida_vol"], col)),
            "saida_inv": int(read_val(ws, HC_ROWS["saida_inv"], col)),
            "final": int(read_val(ws, HC_ROWS["final"], col)),
            "turnover": round(read_val(ws, HC_ROWS["turnover"], col), 4),
        }

    # Custo por área
    custo_por_area = {}
    areas = []
    for row in range(AREAS_START, AREAS_END + 1):
        area = ws.cell(row=row, column=1).value
        if area and area != "CUSTO TOTAL":
            areas.append((row, area))

    for mes_key, col in MESES_COLS.items():
        area_data = {}
        for row, area in areas:
            area_data[area] = round(read_val(ws, row, col), 2)
        custo_por_area[mes_key] = area_data

    return custo_mensal, hc, custo_por_area, [a[1] for a in areas]


def extract_demografico(wb, areas_list):
    """Extrai dados demográficos AGREGADOS por área (sem dados individuais)."""
    demografico = {}
    hc_por_area = {}
    per_capita = {}

    for i, aba_nome in enumerate(MESES_NOMES):
        mes_key = f"2026-{i+1:02d}"
        if aba_nome not in wb.sheetnames:
            continue

        ws = wb[aba_nome]
        # Coletar por departamento: contagem, soma idade, soma tempo casa, soma custo
        dept_stats = defaultdict(lambda: {"count": 0, "idade_sum": 0, "tc_sum": 0, "custo_sum": 0})

        for row in range(4, ws.max_row + 1):
            dept = ws.cell(row=row, column=3).value
            if not dept:
                continue
            idade = ws.cell(row=row, column=7).value
            tc = ws.cell(row=row, column=8).value
            custo = ws.cell(row=row, column=24).value

            if not isinstance(idade, (int, float)):
                continue

            dept_stats[dept]["count"] += 1
            dept_stats[dept]["idade_sum"] += float(idade) if idade else 0
            dept_stats[dept]["tc_sum"] += float(tc) if tc else 0
            dept_stats[dept]["custo_sum"] += float(custo) if custo else 0

        # Agregar
        total_count = sum(d["count"] for d in dept_stats.values())
        total_idade = sum(d["idade_sum"] for d in dept_stats.values())
        total_tc = sum(d["tc_sum"] for d in dept_stats.values())

        demografico[mes_key] = {
            "idade_media": round(total_idade / total_count, 1) if total_count else 0,
            "tempo_casa_medio": round(total_tc / total_count, 1) if total_count else 0,
            "idade_por_area": {d: round(s["idade_sum"] / s["count"], 1) if s["count"] else 0
                               for d, s in dept_stats.items()},
            "tempo_casa_por_area": {d: round(s["tc_sum"] / s["count"], 1) if s["count"] else 0
                                    for d, s in dept_stats.items()},
        }

        hc_por_area[mes_key] = {d: s["count"] for d, s in dept_stats.items()}
        per_capita[mes_key] = {d: round(s["custo_sum"] / s["count"], 2) if s["count"] else 0
                               for d, s in dept_stats.items()}

    return demografico, hc_por_area, per_capita


import re
from datetime import datetime as _dt

# Colunas da planilha mensal (dados individuais)
# Cols 25-29 (Banco, Código, Pix, Agência, Conta) NUNCA incluídos — LGPD
COL_NOME = 2
COL_DEPT = 3
COL_CARGO = 4
COL_ADMISSAO = 5
COL_IDADE = 7
COL_TEMPO_CASA = 8
COL_SALARIO = 9
COL_COMISSAO = 12
COL_BONUS = 13
COL_RESCISAO = 14
COL_CAJU = 15
COL_VT = 16
COL_ESTAC = 19
COL_CLINICA = 20
COL_GYMPASS = 21
COL_CUSTO_TOTAL = 24

GCP_PROJECT_ID = os.environ.get("GCP_PROJECT_ID", "dashboard-koti-omie")
BQ_DATASET = os.environ.get("BQ_DATASET", "studio_koti")


def extract_funcionarios(wb, status_meses):
    """Extrai dados individuais de cada aba mensal para upload ao BigQuery.

    NÃO inclui: CPF, CNPJ, dados bancários (Pix, Agência, Conta).
    """
    rows = []
    for i, aba_nome in enumerate(MESES_NOMES):
        mes_key = f"2026-{i+1:02d}"
        if aba_nome not in wb.sheetnames:
            continue

        ws = wb[aba_nome]
        status = status_meses.get(mes_key, "projecao")

        for row in range(4, ws.max_row + 1):
            nome = ws.cell(row=row, column=COL_NOME).value
            if not nome or not isinstance(nome, str):
                continue
            if nome.strip().upper() in ("TOTAIS", "TOTAL"):
                continue

            # Pular subtotais de departamento: linhas onde "nome" é o nome
            # do departamento (ex: COMERCIAL, ARQUITETURA) sem cargo individual
            cargo_val = ws.cell(row=row, column=COL_CARGO).value
            dept_val = ws.cell(row=row, column=COL_DEPT).value
            nome_upper = nome.strip().upper()
            # Se não tem cargo E o nome parece um departamento (all caps, sem espaço ou com poucos chars)
            # OU se o nome é igual ao departamento → é subtotal
            if not cargo_val or not str(cargo_val).strip():
                # Sem cargo = provavelmente subtotal de departamento
                if dept_val and str(dept_val).strip().upper() == nome_upper:
                    continue  # nome == departamento → subtotal
                if nome_upper == nome.strip() and " " not in nome.strip():
                    continue  # ALL CAPS sem espaço (ex: "COMERCIAL") → subtotal

            sal = ws.cell(row=row, column=COL_SALARIO).value
            if not isinstance(sal, (int, float)) or sal <= 0:
                continue

            def _num(col):
                v = ws.cell(row=row, column=col).value
                return round(float(v), 2) if isinstance(v, (int, float)) else 0.0

            admissao = ws.cell(row=row, column=COL_ADMISSAO).value
            data_adm = None
            if isinstance(admissao, _dt):
                data_adm = admissao.strftime("%Y-%m-%d")

            benef = _num(COL_CAJU) + _num(COL_VT) + _num(COL_ESTAC) + _num(COL_CLINICA) + _num(COL_GYMPASS)

            # custo_total = coluna 24 da planilha (já inclui encargos CLT para faxineiras)
            # menos rescisão (evento pontual, não custo recorrente)
            salario = _num(COL_SALARIO)
            comissao = _num(COL_COMISSAO)
            bonus = _num(COL_BONUS)
            rescisao = _num(COL_RESCISAO)
            custo = _num(COL_CUSTO_TOTAL) - rescisao

            rows.append({
                "nome": nome.strip(),
                "departamento": (ws.cell(row=row, column=COL_DEPT).value or "").strip(),
                "cargo": (ws.cell(row=row, column=COL_CARGO).value or "").strip(),
                "data_admissao": data_adm,
                "idade": int(_num(COL_IDADE)) if _num(COL_IDADE) > 0 else None,
                "tempo_casa_meses": _num(COL_TEMPO_CASA),
                "salario": salario,
                "comissao": comissao,
                "bonus": bonus,
                "rescisao": rescisao,
                "beneficios": round(benef, 2),
                "custo_total": round(custo, 2),
                "mes_referencia": mes_key,
                "status": status,
            })

    return rows


def upload_to_bq(rows):
    """Faz upload da folha para BigQuery (WRITE_TRUNCATE — full replace)."""
    try:
        from google.cloud import bigquery
    except ImportError:
        print("  ⚠ google-cloud-bigquery não instalado — pulando upload BQ")
        return False

    table_id = f"{GCP_PROJECT_ID}.{BQ_DATASET}.folha_funcionarios"
    client = bigquery.Client(project=GCP_PROJECT_ID)

    job_config = bigquery.LoadJobConfig(
        write_disposition="WRITE_TRUNCATE",
        schema=[
            bigquery.SchemaField("nome", "STRING", mode="REQUIRED"),
            bigquery.SchemaField("departamento", "STRING"),
            bigquery.SchemaField("cargo", "STRING"),
            bigquery.SchemaField("data_admissao", "DATE"),
            bigquery.SchemaField("idade", "INT64"),
            bigquery.SchemaField("tempo_casa_meses", "FLOAT64"),
            bigquery.SchemaField("salario", "FLOAT64"),
            bigquery.SchemaField("comissao", "FLOAT64"),
            bigquery.SchemaField("bonus", "FLOAT64"),
            bigquery.SchemaField("rescisao", "FLOAT64"),
            bigquery.SchemaField("beneficios", "FLOAT64"),
            bigquery.SchemaField("custo_total", "FLOAT64"),
            bigquery.SchemaField("mes_referencia", "STRING", mode="REQUIRED"),
            bigquery.SchemaField("status", "STRING"),
        ],
    )

    job = client.load_table_from_json(rows, table_id, job_config=job_config)
    job.result()  # aguarda
    print(f"  ✅ BigQuery: {job.output_rows} linhas em {table_id}")
    return True


# Classificação de nível hierárquico por regex no nome do cargo
NIVEL_RULES = [
    (re.compile(r"gerente", re.IGNORECASE), "Gerência"),
    (re.compile(r"coordenador", re.IGNORECASE), "Coordenação"),
    (re.compile(r"analista|arquiteto|comprador|orçamentista|controller|fiscal", re.IGNORECASE), "Analista/Especialista"),
    (re.compile(r"assistente|auxiliar|almoxarife", re.IGNORECASE), "Assistente/Operacional"),
]

NIVEL_ORDER = ["Gerência", "Coordenação", "Analista/Especialista", "Assistente/Operacional"]


def _classify_nivel(cargo):
    for pattern, nivel in NIVEL_RULES:
        if pattern.search(cargo):
            return nivel
    return "Analista/Especialista"  # fallback


def extract_faixa_salarial(wb):
    """Extrai faixa salarial AGREGADA — SEM nomes individuais (LGPD).

    Retorna: faixa_por_cargo, faixa_por_nivel, faixa_por_area
    """
    # Usar última aba mensal com dados
    cargos: dict[str, list[float]] = {}
    area_salarios: dict[str, list[float]] = {}

    for aba_nome in reversed(MESES_NOMES):
        if aba_nome not in wb.sheetnames:
            continue
        ws = wb[aba_nome]
        found = False
        for row in range(4, ws.max_row + 1):
            cargo = ws.cell(row=row, column=4).value
            sal = ws.cell(row=row, column=9).value
            dept = ws.cell(row=row, column=3).value
            if not cargo or not isinstance(sal, (int, float)) or sal <= 0:
                continue
            found = True
            cargos.setdefault(cargo, []).append(float(sal))
            if dept:
                area_salarios.setdefault(dept, []).append(float(sal))
        if found:
            break

    def _stats(vals):
        return {
            "min": round(min(vals), 2),
            "media": round(sum(vals) / len(vals), 2),
            "max": round(max(vals), 2),
            "qtd": len(vals),
        }

    # Faixa por cargo (legado)
    faixa = {cargo: _stats(vals) for cargo, vals in cargos.items()}

    # Faixa por nível hierárquico
    nivel_salarios: dict[str, list[float]] = {}
    for cargo, vals in cargos.items():
        nivel = _classify_nivel(cargo)
        nivel_salarios.setdefault(nivel, []).extend(vals)
    faixa_nivel = {}
    for nivel in NIVEL_ORDER:
        if nivel in nivel_salarios:
            faixa_nivel[nivel] = _stats(nivel_salarios[nivel])

    # Faixa por área
    faixa_area = {area: _stats(vals) for area, vals in area_salarios.items()}

    return faixa, faixa_nivel, faixa_area


def project_forward(data_by_month, is_zero_fn=None):
    """Preenche meses zerados com o último mês não-zero anterior.

    Preenche "buracos": se Jun tem dados, Jul-Nov são zero e Dez tem dados,
    Jul-Nov recebem cópia de Jun.
    """
    if not data_by_month:
        return data_by_month

    if is_zero_fn is None:
        def is_zero_fn(d):
            if isinstance(d, dict):
                return d.get("total", sum(d.values())) == 0
            return d == 0

    import copy
    meses_ord = sorted(data_by_month.keys())
    last_nonzero = None

    for mes in meses_ord:
        if is_zero_fn(data_by_month[mes]):
            if last_nonzero is not None:
                data_by_month[mes] = copy.deepcopy(data_by_month[last_nonzero])
        else:
            last_nonzero = mes

    return data_by_month


def project_forward_missing(data_by_month, all_months):
    """Preenche meses ausentes (não presentes no dict) com o último mês existente."""
    if not data_by_month:
        return data_by_month

    import copy
    meses_ord = sorted(all_months)
    last_good = None

    for mes in meses_ord:
        if mes in data_by_month:
            last_good = mes
        elif last_good is not None:
            data_by_month[mes] = copy.deepcopy(data_by_month[last_good])

    return data_by_month


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    filepath = args[0] if args else DEFAULT_PATH

    if not os.path.exists(filepath):
        print(f"⚠ Planilha não encontrada: {filepath}")
        sys.exit(0)

    from openpyxl import load_workbook
    print(f"📊 Lendo planilha: {filepath}")
    wb = load_workbook(filepath, data_only=True)

    status_meses, mes_ref = detect_status(wb)
    custo_mensal, hc, custo_por_area, areas_list = extract_resumo(wb)
    demografico, hc_por_area, per_capita = extract_demografico(wb, areas_list)
    faixa_salarial, faixa_por_nivel, faixa_por_area = extract_faixa_salarial(wb)

    # Composição de custo YTD (meses realizados + andamento)
    composicao_ytd = defaultdict(float)
    for mes_key, status in status_meses.items():
        if status in ("realizado", "andamento"):
            for row, key, label in RUBRICAS:
                composicao_ytd[label] += custo_mensal.get(mes_key, {}).get(key, 0)
    composicao_ytd = {k: round(v, 2) for k, v in composicao_ytd.items() if v > 0}

    # Fluxo de caixa (aba FLUXO DE CAIXA)
    fluxo_caixa = {}
    if "FLUXO DE CAIXA" in wb.sheetnames:
        ws_fc = wb["FLUXO DE CAIXA"]
        fc_rubricas = [
            (4, "salarios"), (5, "caju"), (6, "vale_transporte"),
            (7, "inss_fgts"), (8, "clinica"), (9, "gympass"),
            (10, "estacionamento"), (11, "rescisao"), (12, "comissao"),
        ]
        for mes_key, col in MESES_COLS.items():
            fc = {}
            for row, key in fc_rubricas:
                fc[key] = read_val(ws_fc, row, col)
            fc["total"] = read_val(ws_fc, 13, col)
            fluxo_caixa[mes_key] = fc

    # --- Projeção: preencher meses vazios com último mês com dados ---
    all_months = sorted(MESES_COLS.keys())

    project_forward(custo_mensal)
    project_forward(custo_por_area,
                    is_zero_fn=lambda d: isinstance(d, dict) and sum(d.values()) == 0)
    project_forward(fluxo_caixa)

    # demografico, hc_por_area, per_capita só existem para meses com aba na planilha
    project_forward_missing(demografico, all_months)
    project_forward_missing(hc_por_area, all_months)
    project_forward_missing(per_capita, all_months)

    n_projected = sum(1 for m in all_months
                      if status_meses.get(m) == "projecao"
                      and custo_mensal.get(m, {}).get("total", 0) > 0)
    if n_projected:
        print(f"  📈 {n_projected} meses projetados (último dado real replicado)")

    # Headcount histórico (hardcoded 2024-2025, 2026 do RESUMO)
    hc_2026 = [hc.get(f"2026-{m:02d}", {}).get("final", 0) for m in range(1, 13)]
    historico_hc = {
        "2024": [26, 29, 33, 32, 29, 32, 35, 36, 35, 34, 35, 34],
        "2025": [30, 34, 37, 37, 41, 42, 43, 43, 34, 34, 33, 32],
        "2026": hc_2026,
    }

    result = {
        "mes_referencia": mes_ref,
        "status_meses": status_meses,
        "custo_mensal": custo_mensal,
        "hc": hc,
        "custo_por_area": custo_por_area,
        "hc_por_area": hc_por_area,
        "per_capita_area": per_capita,
        "demografico": demografico,
        "composicao_custo_ytd": composicao_ytd,
        "fluxo_caixa": fluxo_caixa,
        "historico_hc": historico_hc,
        "faixa_salarial": faixa_salarial,
        "faixa_por_nivel": faixa_por_nivel,
        "faixa_por_area": faixa_por_area,
    }

    output = os.path.join(SCRIPT_DIR, "rh_data.json")
    with open(output, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"  ✅ {len(custo_mensal)} meses extraídos")
    print(f"  ✅ {len(areas_list)} áreas: {', '.join(areas_list)}")
    print(f"  ✅ Salvo em: {output} ({os.path.getsize(output) / 1024:.1f} KB)")
    print(f"  🔒 Sem dados pessoais no JSON (LGPD)")

    # Upload dados individuais para BigQuery (tabela restrita)
    if "--no-bq" not in sys.argv:
        funcionarios = extract_funcionarios(wb, status_meses)
        if funcionarios:
            print(f"  📤 Enviando {len(funcionarios)} registros para BigQuery...")
            try:
                upload_to_bq(funcionarios)
            except Exception as e:
                print(f"  ⚠ Erro no upload BQ: {e}")
    else:
        print(f"  ⏭ Upload BQ pulado (--no-bq)")


if __name__ == "__main__":
    main()
